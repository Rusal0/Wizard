import streamlit as st
import pandas as pd
from io import BytesIO
import zipfile
import openpyxl
from openpyxl import load_workbook
import hashlib

st.title('Excel Wizard')


# Function to split an Excel file into separate files for each sheet with formatting
def split_excel(file):
  original_wb = load_workbook(file, data_only=True)

  output = BytesIO()
  with zipfile.ZipFile(output, 'w') as zf:
    for sheet_name in original_wb.sheetnames:
      new_wb = openpyxl.Workbook()
      new_sheet = new_wb.active
      new_sheet.title = sheet_name

      original_sheet = original_wb[sheet_name]

      for row in original_sheet.iter_rows():
        for cell in row:
          new_cell = new_sheet[cell.coordinate]
          new_cell.value = cell.value

          # Copy all formatting attributes
          new_cell.font = cell.font.copy()
          new_cell.border = cell.border.copy()
          new_cell.alignment = cell.alignment.copy()
          new_cell.fill = cell.fill.copy()
          new_cell.number_format = cell.number_format

      with BytesIO() as sheet_output:
        new_wb.save(sheet_output)
        zf.writestr(f"{sheet_name}.xlsx", sheet_output.getvalue())

  output.seek(0)
  return output


# Function to merge (union) multiple Excel files into separate sheets within one Excel file
def merge_excels(files):
  combined_output = BytesIO()

  with pd.ExcelWriter(combined_output, engine='xlsxwriter') as writer:
    for file in files:
      try:
        excel_data = pd.read_excel(file, engine='openpyxl')

        # Check if excel_data is a DataFrame
        if isinstance(excel_data, pd.DataFrame):
          sheet_name = file.name.split('.')[0]  # Extract the file name without extension
          excel_data.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
          for sheet_name in excel_data.sheetnames:
            sheet_data = excel_data[sheet_name]

            # Generate a unique sheet name using a hash for potential duplicate sheet names
            with BytesIO(sheet_data.read()) as f:
              file_data = f.read()
              unique_id = hashlib.sha1(file_data + sheet_name.encode()).hexdigest()[:10]
              new_sheet_name = f"{unique_id}_{sheet_name}"

            sheet_data.to_excel(writer, sheet_name=new_sheet_name, index=False)

      except Exception as e:
        st.error(f"Error processing file {file.name}: {str(e)}")

  combined_output.seek(0)
  return combined_output


# File upload options
st.sidebar.title("Excel Wizard Options")
option = st.sidebar.radio("Choose an action", ('Split Excel by Sheets', 'Merge Excel Files'))

# Split Excel File
if option == 'Split Excel by Sheets':
  uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])
  if uploaded_file is not None:
    # Use st.empty() to create a space for the loading icon
    loading_icon = st.empty()
    loading_icon.markdown("**Processing...** (Please wait)")

    split_result = split_excel(uploaded_file)
    loading_icon.empty()  # Clear the loading icon after processing

    # Add the download button after processing is complete
    download_button = st.download_button("Download Split Files (ZIP)", data=split_result, file_name="split_sheets.zip")

# Merge Excel Files
elif option == 'Merge Excel Files':
  uploaded_files = st.file_uploader("Upload multiple Excel files", type=["xlsx"], accept_multiple_files=True)
  if uploaded_files:
    # Use st.empty() to create a space for the loading icon
    loading_icon = st.empty()
    loading_icon.markdown("**Processing...** (Please wait)")

    merged_result = merge_excels(uploaded_files)
    loading_icon.empty()  # Clear the loading icon after processing

    # Add the download button after processing is complete
    download_button = st.download_button("Download Merged File", data=merged_result, file_name="merged_file.xlsx")
